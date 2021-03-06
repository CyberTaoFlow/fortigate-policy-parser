#!/usr/local/bin/python

import argparse
import re
import pprint
import openpyxl

pp = pprint.PrettyPrinter(indent=2)

parser = argparse.ArgumentParser(description="Process Fortigate configuration \
    and create communication matrix excel sheet.")
parser.add_argument('-f', action='store',
                   metavar='<configuration-file>',
                   help='path to configuration file',
                   required=True)

parser.add_argument('-o', action='store',
                    metavar='<output-file>.xls',
                    help='path to configuration file',
                    required=False)

args = parser.parse_args()
CONFIGFILE = vars(args)['f']
print(("Parsing Configuration File: %s" % CONFIGFILE))

try:
    fullconfigstr = open(CONFIGFILE, 'r').read()
except:
    print(("Error reading config file: %s" % CONFIGFILE))

fullconfiglines = fullconfigstr.splitlines()

allpolicies = fullconfiglines[fullconfiglines.index('config firewall policy')+1:]
allpolicies = allpolicies[:allpolicies.index('end')]

policydict = dict()

for line in allpolicies:
    try:
        if line.strip().startswith('edit'):
            policyid = re.match(r'edit (\d*)', line.strip()).groups()[0]
            policydict[policyid] = dict()
        elif line.strip() != 'next' and line.strip().startswith('set'):
            key, val = re.match(r'^set (\S*) (.+)$', line.strip()).groups()
            policydict[policyid][key] = val
    except:
        print(("Error on line: %s" % line))
        raise

print(("Total policies: %d" % len(list(policydict.keys()))))

if 'o' in vars(args):
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title = "Policies"
    row = 2
    columns = list()
    columns.append('id')
    for pid in list(policydict.keys()):
        for key in policydict[pid]:
            if not key in columns:
                columns.append(key)
            if key == 'service' or key == 'srcaddr' or key == 'dstaddr':
                policydict[pid][key] = policydict[pid][key].replace('" "', '","').replace('"', '')
                if key == 'service':
                    policydict[pid][key] = policydict[pid][key].replace('_', '-')
                policydict[pid][key] = policydict[pid][key].split(',')

    for pid in list(policydict.keys()):
        sheet.cell(row=row, column=columns.index('id') + 1).value = pid
        for key in policydict[pid]:
            if type(policydict[pid][key]) == type(list()):
                val = ',\n'.join(policydict[pid][key])
                sheet.cell(row=row, column=columns.index(key)+1).value = val
            else:
                sheet.cell(row=row, column=columns.index(key)+1).value = policydict[pid][key]
        row += 1
        pp.pprint(policydict[pid])
    # Write the header
    for item in columns:
        sheet.cell(row=1, column=columns.index(item) + 1).value = item
    wb.save(vars(args)['o'])